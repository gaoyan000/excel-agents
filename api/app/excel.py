"""Excel ingestion (.xlsx + .xls) -> CSV with header detection + total filter.

Why this module exists
----------------------
Real-world spreadsheets (especially the China-first client's logistics
sheets) routinely contain:

  - A merged TITLE row above the actual column headers
        e.g. "三利物流发货费用(2023.1月)" spanning A1:O1
  - SUBTOTAL / GRAND-TOTAL rows interleaved with data
        e.g. "一汽红旗 汇总 | ... | 1449.40"
              "总计      | ... | 3563.60"

DuckDB's native `read_xlsx` assumes row 1 is the header and has no notion
of "this row is a subtotal, ignore it". Both quirks would silently corrupt
ingestion (the bilingual heuristic would see "三利物流..." as a column
name and SUM(revenue) would double-count subtotals).

Strategy: pre-process every Excel file in Python before DuckDB ever sees
it. Read all cells into list[list[str]]; pick the real header row by
heuristic; drop everything above it; filter out total rows; write the
result as CSV. Downstream `_scan_expr` then reads via `read_csv_auto`,
the same code path that handles plain CSV uploads.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime


# --- text-shaped date detection --------------------------------------------
#
# Real-world spreadsheets carry "dates" that aren't actually date cells:
# the column was formatted as Text or General, and the user typed
# "2023/3/15" or "2023年3月15日". openpyxl returns these as plain strings,
# DuckDB's TRY_CAST(... AS DATE) requires YYYY-MM-DD so it returns NULL,
# and downstream month(date_col) IS NULL produces stray '月' columns in
# pivots. Catch these at ingest time and rewrite to ISO so the rest of
# the pipeline sees clean dates.
_DATE_SEP = re.compile(
    r"^\s*(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})"
    r"(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?\s*$"
)
_DATE_ZH = re.compile(
    r"^\s*(\d{4})年(\d{1,2})月(\d{1,2})日\s*$"
)


def _maybe_iso_date(s: str) -> str | None:
    """Detect a date-shaped text and return ISO (YYYY-MM-DDTHH:MM:SS) or
    None when the string isn't a date."""
    m = _DATE_SEP.match(s)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if m.group(4) is not None:
                hh = int(m.group(4))
                mm = int(m.group(5))
                ss = int(m.group(6) or 0)
                return datetime(y, mo, d, hh, mm, ss).isoformat()
            return datetime(y, mo, d).isoformat()
        except ValueError:
            return None
    m = _DATE_ZH.match(s)
    if m:
        try:
            return datetime(
                int(m.group(1)), int(m.group(2)), int(m.group(3))
            ).isoformat()
        except ValueError:
            return None
    return None


# --- cell -> string normalization ------------------------------------------

def _cell_to_str(v) -> str:
    """Render an openpyxl cell value as the string we write to CSV.

    openpyxl returns Python natives: None, str, int, float, datetime,
    bool. Dates become ISO strings (not Excel serials); integer-valued
    floats become ints (no trailing .0); date-shaped text strings get
    rewritten to ISO so downstream TRY_CAST(... AS DATE) succeeds.
    """
    if v is None:
        return ""
    if hasattr(v, "isoformat"):  # datetime, date, time
        return v.isoformat()
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, str):
        iso = _maybe_iso_date(v)
        return iso if iso is not None else v
    return str(v)


def _xls_cell_to_str(book, cell) -> str:
    """xlrd version of _cell_to_str. Same target output, including the
    text-date detection for cells stored as Text but typed as dates."""
    import xlrd

    ct = cell.ctype
    if ct in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return ""
    if ct == xlrd.XL_CELL_TEXT:
        iso = _maybe_iso_date(cell.value)
        return iso if iso is not None else cell.value
    if ct == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if v.is_integer() else str(v)
    if ct == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, book.datemode).isoformat()
    if ct == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    return ""


# --- format-specific readers -----------------------------------------------

def _read_xlsx_rows(raw: bytes) -> list[list[str]]:
    """First sheet -> rectangular grid of strings.

    read_only=False (the slower mode) so openpyxl auto-resolves date-
    formatted cells: a cell formatted as 'yyyy-mm-dd' with serial value
    44197 comes back as datetime(2021,1,1) instead of the raw float
    44197.0 — which would otherwise hit downstream as "44197" and fail
    TRY_CAST(... AS DATE). read_only=True occasionally misses these
    conversions on non-standard custom formats.

    data_only=True keeps formulas resolved to their cached values, so a
    cell with =TODAY() reads as a real datetime, not the formula text.

    Trade-off: ~3-5x slower than read_only on large sheets. For the
    typical spreadsheet sizes seen in this product (hundreds to a few
    thousand rows) the wall-clock cost is negligible.
    """
    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(raw), read_only=False, data_only=True
    )
    ws = wb[wb.sheetnames[0]]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_str(v) for v in row])
    return rows


def _read_xls_rows(raw: bytes) -> list[list[str]]:
    import xlrd  # lazy: only the .xls path needs it

    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    rows: list[list[str]] = []
    for r in range(sheet.nrows):
        rows.append([
            _xls_cell_to_str(book, sheet.cell(r, c))
            for c in range(sheet.ncols)
        ])
    return rows


# --- header detection ------------------------------------------------------

def _is_numeric(s: str) -> bool:
    s = s.strip().replace(",", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _score_header_row(
    row: list[str], nxt: list[str] | None, ncols: int
) -> float:
    """Score a row's likelihood of being the column header.

    A header row is: dense (most cells non-empty), label-like (short,
    non-numeric, mostly distinct cells), and ideally followed by a row
    that contains numeric data.
    """
    nonempty = [c for c in row if c.strip()]
    if not nonempty:
        return -10.0  # blank row never wins

    score = 0.0
    # Density — title rows are sparse, headers are dense.
    score += 2.0 * (len(nonempty) / ncols)
    # Title-row pattern: a single non-empty cell across the row.
    if len(nonempty) == 1:
        score -= 2.0
    # Label-likeness: short cells without numeric content.
    short = sum(1 for c in nonempty if len(c.strip()) <= 30)
    numeric = sum(1 for c in nonempty if _is_numeric(c))
    score += 1.5 * (short / len(nonempty))
    score -= 2.0 * (numeric / len(nonempty))
    # Distinctness: labels are mostly unique; repeated cells suggest data.
    distinct = len({c.strip() for c in nonempty})
    score += 1.0 * (distinct / len(nonempty))
    # Bonus: the next row carries numeric values (a header precedes data).
    if nxt is not None:
        nxt_nonempty = [c for c in nxt if c.strip()]
        if nxt_nonempty and sum(
            1 for c in nxt_nonempty if _is_numeric(c)
        ) >= 2:
            score += 1.0
    return score


def detect_header_row(rows: list[list[str]]) -> int:
    """Return the index of the most likely header row within rows[:10].

    Defaults to 0 (the historical / CSV-style behavior) when no row
    out-scores it — so plain spreadsheets without a title row are
    unaffected.
    """
    if not rows:
        return 0
    ncols = max((len(r) for r in rows[:20]), default=1)
    scan_until = min(10, len(rows))
    best_idx = 0
    best_score = _score_header_row(
        rows[0], rows[1] if len(rows) > 1 else None, ncols
    )
    for i in range(1, scan_until):
        nxt = rows[i + 1] if i + 1 < len(rows) else None
        s = _score_header_row(rows[i], nxt, ncols)
        if s > best_score:
            best_score, best_idx = s, i
    return best_idx


# --- total-row filter ------------------------------------------------------

_TOTAL_MARKERS_ZH = ("汇总", "合计", "小计", "总计")
_TOTAL_MARKERS_EN = ("total", "subtotal", "grand total", "sum")


def is_total_row(row: list[str]) -> bool:
    """True if `row` looks like a subtotal / grand-total row.

    Heuristic: very few non-empty cells (typically a label + a sum) AND
    at least one of those cells carries a total marker. The "mostly
    empty" guard prevents mis-flagging the header row, which contains
    things like 运费合计 (column name with 合计 as a substring) but is
    dense (15+ non-empty cells), so it never passes this filter.
    """
    nonempty = [c.strip() for c in row if c.strip()]
    if not nonempty or len(nonempty) > 3:
        return False
    for c in nonempty:
        if any(m in c for m in _TOTAL_MARKERS_ZH):
            return True
        cl = c.lower()
        if cl in _TOTAL_MARKERS_EN:
            return True
    return False


# --- public entry point ----------------------------------------------------

def excel_to_csv(filename: str, raw: bytes) -> tuple[str, bytes] | None:
    """Convert an .xls / .xlsx upload to CSV bytes ready for DuckDB.

    Returns (stored_filename, csv_bytes), where stored_filename is
    `<original>.csv` so the downstream `_scan_expr` routes it through
    `read_csv_auto`. Returns None when `filename` is not an Excel file
    (caller leaves the upload untouched — CSV passes through as-is).
    """
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        rows = _read_xlsx_rows(raw)
    elif lower.endswith(".xls"):
        rows = _read_xls_rows(raw)
    else:
        return None

    if not rows:
        return filename + ".csv", b""

    header_idx = detect_header_row(rows)
    header = rows[header_idx]
    width = len(header)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for row in rows[header_idx + 1:]:
        if is_total_row(row):
            continue
        # Pad / truncate to header width — sheets can have trailing or
        # ragged empty cells; CSV needs rectangular rows.
        if len(row) < width:
            row = row + [""] * (width - len(row))
        elif len(row) > width:
            row = row[:width]
        # Skip completely blank data rows (common in xlsx between sections).
        if not any(c.strip() for c in row):
            continue
        writer.writerow(row)
    return filename + ".csv", buf.getvalue().encode("utf-8")
